import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import Workbook
import os

# Data generation logic
data = [
    ["07 Jan (Wed)", "Initial project kickoff; discussed \"Greenergy\" concept with Prof. Kadam.", "Project Initiation", ""],
    ["09 Jan (Fri)", "Finalized Tech Stack: React 18, Vite, and TailwindCSS for the frontend.", "Planning & Setup", ""],
    ["14 Jan (Wed)", "Minor Blocker: Evaluated Redux vs. Context API for global state. Decided on Redux Toolkit.", "State Management", ""],
    ["16 Jan (Fri)", "Requirement analysis: Identified core features like AQI metrics and Heatmaps.", "Requirements Analysis", ""],
    ["21 Jan (Wed)", "Wireframing initial UI on Figma; planned responsive dashboard layout.", "UI/UX Design", ""],
    ["23 Jan (Fri)", "Setup project directory structure, installed dependencies, and configured Tailwind.", "Project Foundation", ""],
    ["28 Jan (Wed)", "Implemented global styling, typography, and primary color palette.", "Design System", ""],
    ["30 Jan (Fri)", "Created reusable Layout components (Sidebar, Navbar) using Framer Motion.", "UI Components", ""],
    ["04 Feb (Wed)", "Built the Landing Page with mission statement and project overview.", "Frontend Dev", ""],
    ["06 Feb (Fri)", "Set up React Router v6 for navigation between Dashboard and Heatmap pages.", "Routing", ""],
    ["11 Feb (Wed)", "Delay: Researched D3.js vs. Recharts for the heatmap implementation.", "Research", ""],
    ["13 Feb (Fri)", "Discussion with team regarding API structure for fetching real-time weather data.", "Database/API Planning", ""],
    ["18 Feb (Wed)", "Started development of the Main Dashboard; implemented card-based metric displays.", "Dashboard Module", ""],
    ["20 Feb (Fri)", "Integrated Recharts to show historical AQI trends for selected cities.", "Data Visualization", ""],
    ["25 Feb (Wed)", "Blocker: Faced issues with data fetching during API integration. Debugging async calls.", "API Integration", ""],
    ["27 Feb (Fri)", "Successfully fetched live AQI data; implemented error handling for empty states.", "Backend Connectivity", ""],
    ["04 Mar (Wed)", "Developed the dynamic AQI Heatmap using D3.js; integrated color-coded zones.", "Heatmap Module", ""],
    ["06 Mar (Fri)", "Added pollutant selection feature (PM2.5, NO2, O3) to the heatmap filters.", "Advanced Filtering", ""],
    ["11 Mar (Wed)", "Human Factor: Day spent cleaning up redundant CSS and optimizing asset loading.", "Refactoring", ""],
    ["13 Mar (Fri)", "Integrated health advisory logic based on WHO pollutant standards.", "Health Advisory", ""],
    ["18 Mar (Wed)", "Designed and implemented User Search for specific geographic coordinates.", "Search & Map", ""],
    ["20 Mar (Fri)", "Conducted cross-browser testing and fixed responsive layout break-points.", "Testing & QA", ""],
    ["25 Mar (Wed)", "Ongoing: Debugged the ErrorBoundary.jsx to handle API timeout crashes.", "Error Handling", ""],
    ["27 Mar (Fri)", "Refining UI transitions and adding tooltips for complex data points.", "UI/UX Improvement", ""]
]

# Create Excel writer
file_name = "CEP_Logbook_New_Greenergy.xlsx"
df = pd.DataFrame(data, columns=["Date", "Work Done / Description", "Module / Feature", "Sign of Prof. Archana Kadam"])

writer = pd.ExcelWriter(file_name, engine='openpyxl')
df.to_excel(writer, index=False, sheet_name='CEP Logbook')

# Formatting with openpyxl
workbook = writer.book
worksheet = writer.sheets['CEP Logbook']

# Styles
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# Apply formatting to headers
for col in range(1, 5):
    cell = worksheet.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Apply border and alignment to data rows
for row in range(2, len(data) + 2):
    for col in range(1, 5):
        cell = worksheet.cell(row=row, column=col)
        cell.border = thin_border
        if col != 2:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

# Set Column Widths
worksheet.column_dimensions['A'].width = 15
worksheet.column_dimensions['B'].width = 50
worksheet.column_dimensions['C'].width = 25
worksheet.column_dimensions['D'].width = 25

# Save
writer.close()
print(f"Excel file created: {os.path.abspath(file_name)}")
